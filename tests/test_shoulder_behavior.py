"""Integration tests for shoulder landmark and measurement pipeline."""
import importlib
import pathlib
import sys
import types

import numpy as np
import pytest

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))


# ---------------------------------------------------------------------------
# Minimal Polyscope stub — keeps all ps.register_* calls silent in headless mode
# ---------------------------------------------------------------------------

class _FakeHandle:
    def add_color_quantity(self, *args, **kwargs):
        pass
    def set_enabled(self, *args, **kwargs):
        pass


class _FakePS(types.ModuleType):
    def __init__(self):
        super().__init__("polyscope")
        self._surfaces = {}
        self._points = {}
        self._curves = {}

    def register_surface_mesh(self, name, *args, **kwargs):
        h = _FakeHandle()
        self._surfaces[name] = h
        return h

    def get_surface_mesh(self, name):
        return self._surfaces.setdefault(name, _FakeHandle())

    def has_surface_mesh(self, name):
        return name in self._surfaces

    def remove_surface_mesh(self, name):
        self._surfaces.pop(name, None)

    def register_point_cloud(self, name, *args, **kwargs):
        h = _FakeHandle()
        self._points[name] = h
        return h

    def has_point_cloud(self, name):
        return name in self._points

    def get_point_cloud(self, name):
        return self._points.setdefault(name, _FakeHandle())

    def remove_point_cloud(self, name):
        self._points.pop(name, None)

    def register_curve_network(self, name, *args, **kwargs):
        h = _FakeHandle()
        self._curves[name] = h
        return h

    def has_curve_network(self, name):
        return name in self._curves

    def remove_curve_network(self, name):
        self._curves.pop(name, None)


# Install stub before any import of geometry_backend
_fake_ps = _FakePS()
sys.modules["polyscope"] = _fake_ps
sys.modules.pop("geometry_backend", None)


@pytest.fixture(scope="module")
def loaded_content():
    import geometry_backend
    importlib.reload(geometry_backend)          # pick up stub already in sys.modules
    vc = geometry_backend.VisContent()
    sid = list(vc.catalog.subjects.keys())[0]
    vc.load_sizestream(sid)
    return vc


def test_compute_derived_landmarks(loaded_content):
    vc = loaded_content
    vc.compute_derived_landmarks()
    assert len(vc.derived_lm_dict) == 14
    for name in ["NeckFrontLeft", "NeckFrontRight", "NeckBackLeft", "NeckBackRight",
                 "ArmholeDepthFrontLeft", "ArmholeDepthBackLeft",
                 "ArmholeDepthFrontRight", "ArmholeDepthBackRight",
                 "WaistDartFrontLeft", "WaistDartFrontRight",
                 "WaistDartBackLeft", "WaistDartBackRight",
                 "WaistDartUpperBackLeft", "WaistDartUpperBackRight"]:
        assert name in vc.derived_lm_dict
        pos = vc.derived_lm_dict[name]["position"]
        assert pos.shape == (3,)
        assert 0 < pos[1] < 2500   # Y is height in mm; armhole landmarks lie within human body range


def test_compute_shoulder_measurements(loaded_content):
    vc = loaded_content
    if not vc.derived_lm_dict:
        vc.compute_derived_landmarks()
    vc.compute_shoulder_measurements()
    # 8 neck geodesics + 4 shoulder geodesics + 4 shoulder Y projections = 16
    assert len(vc.measurement_results) >= 16
    geo_results = [r for r in vc.measurement_results if r.method == "geodesic"]
    y_results = [r for r in vc.measurement_results if r.method == "y_projection"]
    assert len(geo_results) == 12  # 8 neck + 4 shoulder
    assert len(y_results) == 4
    for r in geo_results:
        assert r.value_mm > 0
        assert r.value_mm < 1000


def test_export_to_excel(loaded_content, tmp_path):
    vc = loaded_content
    if not vc.derived_lm_dict:
        vc.compute_derived_landmarks()
    if not vc.measurement_results:
        vc.compute_shoulder_measurements()
    out = str(tmp_path / "test_export.xlsx")
    vc.export_results_to_excel(out)
    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert "Landmarks" in wb.sheetnames
    assert "Measurements" in wb.sheetnames
    ws_lm = wb["Landmarks"]
    assert ws_lm.max_row > 5
    ws_m = wb["Measurements"]
    assert ws_m.max_row > 1
